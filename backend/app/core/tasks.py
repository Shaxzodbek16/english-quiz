import os
from typing import List
import asyncio
import threading
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import uuid
import time

from app.bot.utils.send_msg2admins import send_xlsx
from app.core.settings import get_settings, Settings
import psycopg2
from psycopg2.extras import RealDictCursor
from app.server.celery_worker import app as celeryapp

celeryapp.loop = asyncio.get_event_loop()
celeryapp.loop_runner = threading.Thread(
    target=celeryapp.loop.run_forever,
    daemon=True,
)
celeryapp.loop_runner.start()

settings: Settings = get_settings()


def get_users_sync(page: int, size: int) -> List[dict]:
    try:
        conn = psycopg2.connect(
            dbname=settings.POSTGRES_DATABASE,
            user=settings.POSTGRES_USER,
            password=settings.POSTGRES_PASSWORD,
            host=settings.POSTGRES_HOST,
            cursor_factory=RealDictCursor,
        )

        with conn.cursor() as cursor:
            offset = page * size
            cursor.execute(
                """
                SELECT first_name,
                       last_name,
                       telegram_id,
                       profile_picture, language, is_active, id, created_at, updated_at
                FROM users
                    LIMIT %s
                OFFSET %s
                """,
                (size, offset),
            )
            results = cursor.fetchall()

        conn.close()
        return [dict(r) for r in results]
    except Exception as e:
        print(f"Database connection error: {e}")
        return []


def make_xlsx(tg_id: int) -> str:
    unique_id = uuid.uuid4().hex[:8]
    filename = f"Users_information_{tg_id}_{unique_id}.xlsx"
    export_dir = "media/excel"
    os.makedirs(export_dir, exist_ok=True)
    file_path = os.path.join(export_dir, filename)

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(
        fill_type="solid", start_color="1F4E78", end_color="1F4E78"
    )
    header_font = Font(bold=True, color="FFFFFF")
    even_row_fill = PatternFill(
        fill_type="solid", start_color="D9D9D9", end_color="D9D9D9"
    )
    custom_headers = [
        "🆔 ID",
        "👤 First Name",
        "👥 Last Name",
        "📱 Telegram ID",
        "✅ Is Active",
        "🖼️ Profile Picture",
        "🌐 Language",
        "📅 Joined Date",
        "⏰ Updated Date",
    ]

    page = 0
    size = 10000
    sheet_index = 1

    while True:
        users = get_users_sync(page, size)
        if not users:
            break

        ws = wb.create_sheet(title=f"Users_{sheet_index}")

        for col_idx, header in enumerate(custom_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, user_data in enumerate(users, 2):
            row_fill = even_row_fill if row_idx % 2 == 0 else None

            created_at = user_data.get("created_at")
            updated_at = user_data.get("updated_at")

            if created_at and hasattr(created_at, "tzinfo") and created_at.tzinfo:
                created_at = created_at.replace(tzinfo=None)
            if updated_at and hasattr(updated_at, "tzinfo") and updated_at.tzinfo:
                updated_at = updated_at.replace(tzinfo=None)

            id_cell = ws.cell(row=row_idx, column=1, value=user_data.get("id"))
            first_name_cell = ws.cell(
                row=row_idx, column=2, value=user_data.get("first_name")
            )
            last_name_cell = ws.cell(
                row=row_idx, column=3, value=user_data.get("last_name")
            )
            telegram_id_cell = ws.cell(
                row=row_idx, column=4, value=user_data.get("telegram_id")
            )

            if row_fill:
                id_cell.fill = row_fill
                first_name_cell.fill = row_fill
                last_name_cell.fill = row_fill
                telegram_id_cell.fill = row_fill

            is_active = user_data.get("is_active", False)
            status_cell = ws.cell(
                row=row_idx, column=5, value="✅" if is_active else "❌"
            )
            if row_fill:
                status_cell.fill = row_fill
            status_cell.alignment = Alignment(horizontal="center")
            status_cell.font = Font(
                bold=True, color="008000" if is_active else "FF0000"
            )

            profile_pic = user_data.get("profile_picture", "")
            pic_cell = ws.cell(row=row_idx, column=6, value=profile_pic)
            if row_fill:
                pic_cell.fill = row_fill
            if profile_pic:
                pic_cell.hyperlink = profile_pic
                pic_cell.font = Font(color="0563C1", underline="single")

            lang_cell = ws.cell(row=row_idx, column=7, value=user_data.get("language"))
            if row_fill:
                lang_cell.fill = row_fill

            joined_cell = ws.cell(row=row_idx, column=8, value=created_at)
            if row_fill:
                joined_cell.fill = row_fill

            updated_cell = ws.cell(row=row_idx, column=9, value=updated_at)
            if row_fill:
                updated_cell.fill = row_fill

        for col_idx in range(1, len(custom_headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15  # Set a fixed width

        note_row = len(users) + 4
        ws.cell(row=note_row, column=1, value="Note:").font = Font(bold=True)
        ws.cell(
            row=note_row,
            column=2,
            value="Click on profile picture URLs to view images in your browser",
        )

        page += 1
        sheet_index += 1

    wb.save(file_path)
    return file_path


@celeryapp.task(name="app.core.tasks.make_file_and_send")
def make_file_and_send(tg_id: int) -> bool:
    try:
        time.sleep(1)
        file_path = make_xlsx(tg_id=tg_id)
        success = False
        try:
            coro = send_xlsx(tg_id=tg_id, file_path=file_path)
            asyncio.run_coroutine_threadsafe(
                coro=coro,
                loop=celeryapp.loop,
            )
            if os.path.exists(file_path):
                os.remove(file_path)
            success = True
        except Exception as e:
            if os.path.exists(file_path):
                os.remove(file_path)
            print(f"Error sending file: {e}")
        return success
    except Exception as e:
        print(f"Error creating or sending file: {e}")
        return False
